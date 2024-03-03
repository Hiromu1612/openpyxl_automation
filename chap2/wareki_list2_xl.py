import openpyxl

book=openpyxl.Workbook()
sheet=book.active

sheet["A1"]="西暦"
sheet["B1"]="和暦"

start_y=1930
for i in range(100):
    seireki=str(start_y+i)
    wareki='=TEXT("{}/1/1","gggg年")'.format(seireki) #format関数は文字列内の{}を変数に置き換える

    cell=sheet.cell(row=i+2,column=1,value=seireki+"年")
    cell=sheet.cell(row=i+2,column=2,value=wareki)

book.save('chap2/excell_file/wareki_list2_xl.xlsx')