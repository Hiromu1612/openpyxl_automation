import openpyxl

book=openpyxl.load_workbook(r"C:\Users\1612h\Downloads\0src\samples\ch2\all-customer.xlsx")
#名簿のシートを選択
sheet=book["名簿"]
customers=[["名前","住所","購入プラン"]]

for row in sheet.iter_rows(min_row=3): #max_rowを指定しないと最終行まで取得
    r=[]
    for cell in row:
        r.append(cell.value)
    if r[0] is None:#空白のセルに到達したら終了
        break
    customers_area=r[1]
    if customers_area=="横浜市" or customers_area=="名古屋市":#===は等しいかどうかを判定する演算子
        customers.append(r)
        print(r)

#新規ブックを作成
new_book=openpyxl.Workbook()
new_sheet=new_book.active
new_sheet["A1"]="横浜と名古屋の顧客名簿"

#新規ブックにデータを書き込む
for i in range(len(customers)):
    for j in range(len(customers[i])):
        new_sheet.cell(row=i+2,column=j+1,value=customers[i][j]) #[i][j]の値を書き込む
new_book.save("chap2/excell_file/yokohama-nagoya-customer.xlsx")#新規ブックを保存