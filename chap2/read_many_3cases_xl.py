import openpyxl

book=openpyxl.load_workbook('chap2/excell_file/coordinate_xl.xlsx')
sheet=book.active

#1.連続でセルの値を取得
for i in range(2,5):
    r=[]
    for j in range(2,5):
        v=sheet.cell(row=j,column=i).value
        r.append(v) #append()でリストに要素を追加
    print(r)

#2.特定の範囲のセルの値を取得
for row in sheet["B2":"D4"]:#左上のセルと右下のセルを指定
    r=[]
    for cell in row:
        r.append(cell.value)
    print(r)

#3.iter_rows()を使ってセルの値を取得
it=sheet.iter_rows(min_row=2,max_row=4,min_col=2,max_col=4)
for row in it:
    for cell in row:
        print(cell.value)