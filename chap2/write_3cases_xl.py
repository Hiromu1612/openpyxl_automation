import openpyxl

book=openpyxl.Workbook() #新規のワークブックを作成
sheet=book.active #アクティブ(最後に開いた)なワークシートを取得

sheet['A1']='Hello World!'

sheet.cell(row=2,column=1,value="Hello World!")#連続で値を書き込む、読み込む場合はこの方法が便利

cell=sheet.cell(row=3,column=1)
cell.value="Hello World!"

book.save('chap2/write_3cases_xl.xlsx') #chap2フォルダにhello.xlsxとして保存