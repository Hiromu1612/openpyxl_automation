import openpyxl
book=openpyxl.Workbook()
sheet=book.active

#横幅を指定
sheet.column_dimensions["B"].width=40 #B列の横幅を40に指定
#縦幅を指定
sheet.row_dimensions[2].height=40 #2行目の縦幅を40に指定

cell=sheet["B2"]
cell.value="喜びにあふれた心は良い薬となる"

#テキスト配置の指定
from openpyxl.styles import Alignment #alignment(配置)
cell.alignment=Alignment(horizontal="center",vertical="center") #水平方向は中央、垂直方向は中央に配置

#罫線の指定
from openpyxl.styles.borders import Border,Side #border(罫線),side(線)
cell.border=Border(top=Side(style="thin",color="000000"),bottom=Side(style="thin"),left=Side(style="thin"),right=Side(style="thin")) #上下左右に細い罫線を指定

#フォントの指定
from openpyxl.styles import Font
cell.font=Font(name="ＭＳ Ｐゴシック",size=16,bold=True,italic=True,color="FFFFFF") #フォントをMS Pゴシック、サイズを16、太字、イタリックに指定

#セルの背景色の指定
from openpyxl.styles import PatternFill
cell.fill=PatternFill(patternType="solid",fgColor="0000FF") #背景色を青色に指定

book.save("chap2/excell_file/text_style_xl.xlsx")