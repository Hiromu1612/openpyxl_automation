import openpyxl
import datetime

book=openpyxl.Workbook()
sheet=book.active

this_year=datetime.date.today().year

for i in range(80):
    age=i #満年齢
    year=this_year-age #生まれた年

    age_cell=sheet.cell(row=i+1,column=1)
    age_cell.value=str(age)+'歳'
    year_cell=sheet.cell(row=i+1,column=2)
    year_cell.value=str(year)+'年'

book.save('chap2/excell_file/year_list_xl.xlsx')