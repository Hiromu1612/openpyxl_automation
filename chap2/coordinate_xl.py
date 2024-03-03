import openpyxl

book=openpyxl.Workbook() 
sheet=book.active

for i in range(1,101):
    for j in range(1,101):
        sheet.cell(row=i,column=j,value=sheet.cell(i,j).coordinate)#coordinate属性でセルの座標を取得
        #cell=sheet.cell(i,j)
        #cell.value=cell.coordinate

book.save('chap2/coordinate_xl.xlsx')