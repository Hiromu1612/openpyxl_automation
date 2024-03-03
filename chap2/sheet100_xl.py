import openpyxl
import random

#当たりのシートの番号を決める
atari=random.randint(1,19)

book=openpyxl.Workbook()
book.active["B2"]="当たりが書かれたシートを探そう"

for i in range(1,19):
    sheet_name=str(i)+"枚目"
    sheet=book.create_sheet(title=sheet_name)
    #シートに書き込む
    word="はずれ"
    if i ==atari:
        word="当たり"
    for j in range(1,19):
        for k in range(1,19):
            sheet.cell(row=j,column=k,value=word)

book.save("chap2/excell_file/atari.xlsx")
print("ok,atari=",atari)