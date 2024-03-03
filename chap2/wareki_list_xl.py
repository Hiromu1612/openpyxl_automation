import openpyxl

wareki_table=[{'name':'明治','start':1868,'end':1912},
              {'name':'大正','start':1912,'end':1926},
              {'name':'昭和','start':1926,'end':1989},
              {'name':'平成','start':1989,'end':2019},
              {'name':'令和','start':2019,'end':9999}]

#西暦から和暦に変換する関数を定義
def seireki_to_wareki(year):
    for w in wareki_table:
        if w["start"]<=year<w["end"]:
            y=str(year-w["start"]+1)+"年"
            if y=="1年":
                y="元年"
            return w["name"]+y
    return "不明"

book=openpyxl.Workbook()
sheet=book.active

sheet['A1']='西暦'
sheet['B1']='和暦'

#100年分の西暦と和暦を書き込む
start_y=1930
for i in range(100):
    seireki=start_y+i
    wareki=seireki_to_wareki(seireki)

    cell=sheet.cell(row=i+2,column=1)
    cell.value=str(seireki)+"年"
    cell=sheet.cell(row=i+2,column=2)
    cell.value=wareki

book.save('chap2/excell_file/wareki_list_xl.xlsx')