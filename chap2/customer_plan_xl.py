import openpyxl

book=openpyxl.load_workbook(r"C:\Users\1612h\Downloads\0src\samples\ch2\all-customer.xlsx")
sheet=book["名簿"]

#購入プランごとにシートを分ける
for row in sheet.iter_rows(min_row=3):
    r=[]
    for cell in row:
        r.append(cell.value)
    if r[0] is None:
        break
    (name,area,plan)=r #タプルに代入
    sheet_name=plan+"プラン" #sheet_nameにプラン名を代入

    if sheet_name not in book.sheetnames: #該当するシートがあるか sheetnamesでシート名を取得
        to_sheet=book.create_sheet(title=sheet_name) #to_sheetに新規シートを代入 create_sheet()で新規シートを作成
        to_sheet.append(["名前","住所","プラン"])
    else:
        to_sheet=book[sheet_name]
    to_sheet.append([name,area,plan])

book.save("chap2/excell_file/customer_plan.xlsx")