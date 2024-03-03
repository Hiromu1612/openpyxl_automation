import openpyxl,json
#入出力ファイルを指定
input_file="chap3/excell_file/keiribu_matome(2).json"
output_dir="chap3/excell_file"
template_file="chap3/excell_file/invoice.xlsx"
subject="2月分の請求書"

#メイン処理
def create_invoice():
    with open(input_file,"rt") as f: #rtは読み込みモード
        users=json.load(f) #jsonファイルをPythonの辞書型に変換
    #顧客ごとに請求書を作成
    for name,value in users.items():
        create_user_invoice(name,value)

#テンプレートにデータを埋め込む
def create_user_invoice(name,value):
    book=openpyxl.load_workbook(template_file)
    sheet=book.active
    #シートに名前、件名、金額を書き込む
    sheet["B4"]=name
    sheet["C10"]=subject
    sheet["C11"]=value["total"]
    #明細を書き込む
    for i,item in enumerate(value["items"]):
        date,summary,count,price=item
        sheet.cell(row=15+i,column=2,value=date)
        sheet.cell(row=15+i,column=5,value=summary)
        sheet.cell(row=15+i,column=6,value=count)
        sheet.cell(row=15+i,column=7,value=count*price)
    #ファイルに保存
    output_file=output_dir+"/"+name+"様.xlsx"
    book.save(output_file)
    print("saved",output_file)

if __name__=="__main__":
    create_invoice() #メイン処理を実行