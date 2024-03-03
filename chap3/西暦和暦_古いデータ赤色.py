import openpyxl
from openpyxl.styles import PatternFill
import datetime

input_file=r"C:\Users\1612h\Downloads\0src\samples\ch3\date-sample.xlsx"
output_file="chap3/excell_file/wareki_red.xlsx"
cell_old_data=datetime.datetime(2020,1,1)

#メイン処理
def date_red(input_file,output_file):
    book=openpyxl.load_workbook(input_file)
    #全シート
    for sheet in book.worksheets: #worksheetsはシートのリスト
        #全行
        for row in sheet.iter_rows(min_row=4):
            check_cell(row)
    book.save(output_file)

#セルの値が指定の日付より古い場合は赤色にする
def check_cell(row):
    date=row[0].value
    #日付型でなけらばチェックしない
    if type(date) is not datetime.datetime:
        return
    if date<cell_old_data:
        red=PatternFill(patternType="solid",fgColor="FF0000")#solidは塗りつぶし
        #その行のセル全てに赤色を設定
        for cell in row:
            cell.fill=red

if __name__=="__main__":
    date_red(input_file,output_file) #メイン処理を実行