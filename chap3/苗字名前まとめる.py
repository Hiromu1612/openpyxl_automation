import openpyxl
input_file=r"C:\Users\1612h\Downloads\0src\samples\ch3\name2.xlsx"
output_file="chap3/excell_file/name2.xlsx"

input_book=openpyxl.load_workbook(input_file)
input_sheet=input_book.worksheets[0] #最初のシートを取得

output_book=openpyxl.Workbook()
output_sheet=output_book.active

for row in input_sheet.iter_rows():
    myouji=row[0].value
    namae=row[1].value
    #名前を結合して出力
    fullname=myouji+namae
    output_sheet.append([fullname])

output_book.save(output_file)