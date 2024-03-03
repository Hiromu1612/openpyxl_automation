#最初にファイル名を指定して、後で見返しやすくする
template_file=r"C:\Users\1612h\Downloads\0src\samples\ch3\invoice-template.xlsx"
save_file="chap3/excell_file/invoice.xlsx"

#データ
name="田中"
subject="1月分のご請求"
items=[#商品名、数量、単価
    ["リンゴ",3,320],
    ["バナナ",1,210],
    ["メロン",2,1200]
]

#テンプレートを書く
import openpyxl
book=openpyxl.load_workbook(template_file)
sheet=book.active

sheet["B4"]=name
sheet["C10"]=subject

total_price=0
for i,item in enumerate(items): #enumerate(列挙)でインデックスと要素を同時に取得
    name,count,price=item #itemsの要素をname,count,priceに代入
    total_price_item=count*price
    total_price=total_price+total_price_item

    sheet.cell(row=15+i,column=2,value=name)
    sheet.cell(row=15+i,column=5,value=count)
    sheet.cell(row=15+i,column=6,value=price)
    sheet.cell(row=15+i,column=7,value=total_price_item)
sheet["C11"]=total_price

book.save(save_file)