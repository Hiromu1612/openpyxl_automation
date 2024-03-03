import glob
import openpyxl

target_dir="chap3/excell_file/salesbooks" #ディレクトリは一番上の階層からの相対パス
save_file="chap3/excell_file/keiribu_matome.xlsx"

#メイン処理
def read_files():
    #売上一覧を書き込むブックを作成
    book=openpyxl.Workbook()
    main_sheet=book.active
    enumfiles(main_sheet)#ファイルを列挙して読み込む
    book.save(save_file)#読み込んだデータを保存する

#ファイルを列挙する
def enumfiles(main_sheet):
    files=glob.glob(target_dir+"/*.xlsx")
    for file_name in files:
        read_book(file_name,main_sheet)

#エクセルファイルを読み込む
def read_book(file_name,main_sheet):
    print("read:"+file_name)

    book=openpyxl.load_workbook(file_name,data_only=True) #data_only=Trueで数式を計算結果で読み込む
    sheet=book.active

    #シートの内容を読み込む
    for row in sheet.iter_rows(min_row=4):
        values=[col.value for col in row]
        if values[0]==None:
            break
        print(values)
        main_sheet.append(values)

#メイン処理を実行
if __name__=="__main__": #このpythonファイルをモジュールとして利用できるようになる
    read_files()