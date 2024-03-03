import openpyxl
input_file=r"C:\Users\1612h\Downloads\0src\samples\ch3\name1.xlsx"
output_file="chap3/excell_file/name1.xlsx"

book=openpyxl.load_workbook(input_file)
sheet=book.worksheets[0] #最初のシートを取得

output_book=openpyxl.Workbook()
output_sheet=output_book.active

for row in sheet.iter_rows():
    fullname=row[0].value
    myouji,namae=fullname.split(" ")
    output_sheet.append([myouji,namae])

output_book.save(output_file)